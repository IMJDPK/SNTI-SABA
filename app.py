import os
# Add os import if not already present
os.environ['OAUTHLIB_INSECURE_TRANSPORT'] = '1' # Allow HTTP for local testing


import openpyxl
import redis
import re
import logging
import requests
import json
from datetime import datetime, timedelta, timezone
from flask import Flask, request, jsonify, session, Response, make_response, render_template, send_from_directory, url_for, redirect
from flask_session import Session  # type: ignore
from flask_cors import CORS
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet
from google_calendar_utils import create_meet_event, create_meet_event_oauth, OAUTH_CLIENT_SECRET_FILE
from typing import List, Dict, Any, Optional
from google_auth_oauthlib.flow import Flow
from googleapiclient.discovery import build
import pathlib
import pickle
import uuid
import sys

# Add transformers import for pipeline
try:
    from transformers import pipeline
except ImportError:
    pipeline = None
    print("Warning: transformers not installed. Emotion analysis will be disabled.")

# Google Calendar OAuth2 scopes
SCOPES = ['https://www.googleapis.com/auth/calendar.events', 'https://www.googleapis.com/auth/calendar.readonly']

# Load environment variables
load_dotenv()

# Get Gemini API Key
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")
if not GEMINI_API_KEY:
    raise ValueError("❌ Missing GEMINI_API_KEY in .env file")

# Get model name from environment or default to gemini-2.0-flash
GEMINI_MODEL = os.getenv("GEMINI_MODEL", "gemini-2.0-flash")

# Words per minute rate limiting
WORDS_PER_MINUTE = 300
SECONDS_PER_WORD = 60 / WORDS_PER_MINUTE

# Excel file configuration
EXCEL_FOLDER = "client_data"
os.makedirs(EXCEL_FOLDER, exist_ok=True)

# Optionally skip Hugging Face model loading if SKIP_HF_EMOTION is set
if os.getenv("SKIP_HF_EMOTION", "0") == "1":
    logging.info("Skipping Hugging Face emotion classifier initialization due to SKIP_HF_EMOTION=1")
    emotion_classifier = None
else:
    try:
        if pipeline is None:
            logging.warning("Transformers pipeline not available, skipping emotion classifier")
            emotion_classifier = None
        else:
            emotion_classifier = pipeline(
                "text-classification",
                model="j-hartmann/emotion-english-distilroberta-base",
                return_all_scores=True
            )
    except Exception as e:
        logging.error(f"Failed to load emotion classifier: {e}")
        emotion_classifier = None

def analyze_emotion(text: str) -> Dict[str, float]:
    """
    Analyze the emotional content of text using the emotion classifier.
    
    Args:
        text: Input text to analyze
        
    Returns:
        Dictionary mapping emotion labels to confidence scores
    """
    if not emotion_classifier:
        return {"error": "Emotion classifier not initialized"}
        
    try:
        # Get emotion predictions
        results = emotion_classifier(text)[0]
        
        # Convert to dictionary format
        emotions = {item['label']: float(item['score']) for item in results}
        return emotions
        
    except Exception as e:
        logging.error(f"Error in emotion analysis: {e}")
        return {"error": str(e)}
LEADS_DATA_FILE = os.path.join(EXCEL_FOLDER, "leads.json")  # Define leads data file path
DAILY_REPORTS_FOLDER = os.path.join(EXCEL_FOLDER, "daily_lead_reports")  # New folder for daily reports
os.makedirs(DAILY_REPORTS_FOLDER, exist_ok=True)  # Ensure the new folder is created

LOGS_FOLDER = EXCEL_FOLDER  # <-- Add this line so logs endpoints work

# Define client data folder (same as excel folder)
CLIENT_DATA_FOLDER = EXCEL_FOLDER

# In your app.py, set the correct variable name for the client secret file
CLIENT_SECRETS_FILE = OAUTH_CLIENT_SECRET_FILE  # Use the variable from google_calendar_utils.py

# Validate API key and model
def validate_api_key_and_model():
    try:
        response = requests.get(f"https://generativelanguage.googleapis.com/v1beta/models?key={GEMINI_API_KEY}")
        if response.status_code != 200:
            return False, "Invalid or unauthorized GEMINI_API_KEY"
        models = response.json().get("models", [])
        model_names = [model["name"] for model in models]
        if f"models/{GEMINI_MODEL}" not in model_names:
            logging.warning(f"⚠️ Model '{GEMINI_MODEL}' not found in available models: {model_names}")
        return True, None
    except Exception as e:
        return False, f"Failed to validate API key or model: {str(e)}"

is_valid, error_message = validate_api_key_and_model()
if not is_valid:
    raise ValueError(error_message)

# Functions to load and save leads
def load_leads_from_file() -> List[Dict[str, Any]]:
    """Loads leads from the JSON data file."""
    if not os.path.exists(LEADS_DATA_FILE):
        return []
    try:
        with open(LEADS_DATA_FILE, "r", encoding="utf-8") as f:
            data: List[Dict[str, Any]] = json.load(f)
            if isinstance(data, list):  # Basic validation
                return data
            else:
                logging.error(f"Leads data file {LEADS_DATA_FILE} does not contain a JSON list. Initializing with empty list.")
                return []
    except (json.JSONDecodeError, IOError) as e:
        logging.error(f"Error loading leads from {LEADS_DATA_FILE}: {e}. Initializing with empty list.")
        return []

def save_leads_to_file(leads_data: List[Dict[str, Any]]) -> None:
    """Saves leads to the JSON data file."""
    try:
        with open(LEADS_DATA_FILE, "w", encoding="utf-8") as f:
            json.dump(leads_data, f, indent=4)
        logging.info(f"Leads data saved to {LEADS_DATA_FILE}")
    except IOError as e:
        logging.error(f"Error saving leads to {LEADS_DATA_FILE}: {e}")

def save_to_excel(client_data: Dict[str, Any], client_id: str) -> bool:
    """Save collected client data to Excel file"""
    try:
        filename = f"{EXCEL_FOLDER}/client_{client_id}.xlsx"
        wb: Workbook
        ws: Worksheet
        if os.path.exists(filename):
            wb = openpyxl.load_workbook(filename)
            ws = wb.active
            if ws is None:  # Ensure ws is not None
                ws = wb.create_sheet()
        else:
            wb = Workbook()
            ws = wb.active
            if ws is None:  # Ensure ws is not None
                ws = wb.create_sheet()  # Or handle error appropriately
            headers = ["Timestamp", "Field", "Value"]
            ws.append(headers)
            for cell in ws[1]:  # type: ignore
                if cell:
                    cell.font = Font(bold=True)
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        for field, value in client_data.items():
            row: List[Any] = [timestamp, field, value]
            ws.append(row)
        wb.save(filename)
        logging.info(f"Saved client data to {filename}")
        return True
    except Exception as e:
        logging.error(f"Error saving to Excel: {str(e)}")
        return False

def save_conversation(client_id: str, conversation_history: List[Dict[str, Any]]) -> str:
    """Save conversation history to a file, updating existing conversation or creating new one"""
    try:
        # First, check if there's an existing conversation file for this client
        existing_filename = find_existing_conversation_file(client_id)
        
        if existing_filename:
            # Update existing file
            filename = f"{EXCEL_FOLDER}/{existing_filename}"
            logging.info(f"Updating existing conversation file: {filename}")
        else:
            # Create new file with timestamp
            timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
            filename = f"{EXCEL_FOLDER}/conversation_{client_id}_{timestamp}.txt"
            logging.info(f"Creating new conversation file: {filename}")
        
        # Write the complete conversation history
        with open(filename, "w", encoding="utf-8") as f:
            for entry in conversation_history:
                role: str = entry.get("role", "unknown")  # type: ignore
                parts: List[Dict[str, str]] = entry.get("parts", [])  # type: ignore
                for part in parts:
                    text: str = part.get("text", "")  # type: ignore
                    f.write(f"{role}: {text}\n")
        
        # Get just the filename for database storage
        conversation_filename = os.path.basename(filename)
        
        # Extract client info and link to leads database
        phone = extract_phone_from_chat_id(client_id)
        # Extract email from conversation if available
        email = None
        name = None
        
        # Try to extract email and name from conversation
        for entry in conversation_history:
            if entry.get("role") == "user":
                parts = entry.get("parts", [])
                for part in parts:
                    text = part.get("text", "")
                    # Extract email pattern
                    email_match = re.search(r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b', text)
                    if email_match and not email:
                        email = email_match.group(0).lower()
                    
                    # Extract name pattern
                    name_match = re.search(r"(?:my name is|i'm|i am|this is|call me)\s+([a-zA-Z\s]+)", text, re.IGNORECASE)
                    if name_match and not name:
                        potential_name = name_match.group(1).strip().title()
                        if len(potential_name) > 1 and not any(word in potential_name.lower() for word in ['saba', 'ai', 'bot', 'hello', 'help']):
                            name = potential_name
        
        # Link conversation to lead if we have identifying information
        if phone or email:
            link_conversation_to_lead(conversation_filename, phone, email, name)
        
        # Return just the filename (not full path) for database storage
        return conversation_filename
        
    except Exception as e:
        logging.error(f"Error saving conversation history: {str(e)}")
        return ""

# Logging setup
logging.basicConfig(level=logging.DEBUG)

# Flask setup
app = Flask(__name__)
CORS(app, supports_credentials=True)

# Session configuration (for cookie-based clients)
app.config["SECRET_KEY"] = os.urandom(24)
app.config["SESSION_TYPE"] = os.getenv("SESSION_TYPE", "redis")

# Fixed Redis configuration
redis_host = os.getenv("REDIS_HOST", "localhost")
redis_port = int(os.getenv("REDIS_PORT", 6379))
redis_db = int(os.getenv("REDIS_DB", 0))

app.config["SESSION_REDIS"] = redis.Redis(
    host=redis_host,
    port=redis_port,
    db=redis_db
)

app.config["SESSION_PERMANENT"] = False
app.config["SESSION_USE_SIGNER"] = True
app.config["SESSION_COOKIE_SAMESITE"] = os.getenv("SESSION_COOKIE_SAMESITE", "None")
app.config["SESSION_COOKIE_SECURE"] = os.getenv("SESSION_COOKIE_SECURE", "True") == "True"
app.config["PERMANENT_SESSION_LIFETIME"] = int(os.getenv("SESSION_LIFETIME", 86400))  # 24 hours

Session(app)  # type: ignore

# Rate limiting with Redis storage
redis_url = f"redis://{os.getenv('REDIS_HOST', 'localhost')}:{os.getenv('REDIS_PORT', 6379)}/{os.getenv('REDIS_DB', 0)}"
limiter = Limiter(
    key_func=get_remote_address,
    default_limits=["20000 per day", "5000 per hour"],
    storage_uri=redis_url
)

limiter.init_app(app)

# System instruction
INSTRUCTION_FILE = "system_instruction.txt"
HISTORY_FILE = "system_instruction_history.json"

def get_instruction() -> Optional[str]:
    if not os.path.exists(INSTRUCTION_FILE):
        return None
    with open(INSTRUCTION_FILE, "r", encoding="utf-8") as f:
        return f.read()

def save_instruction(text: str) -> None:
    with open(INSTRUCTION_FILE, "w", encoding="utf-8") as f:
        f.write(text)

def get_history() -> List[Dict[str, str]]:
    if not os.path.exists(HISTORY_FILE):
        return []
    with open(HISTORY_FILE, "r", encoding="utf-8") as f:
        history_data: List[Dict[str, str]] = json.load(f)
        return history_data

def save_history(history: List[Dict[str, str]]) -> None:
    with open(HISTORY_FILE, "w", encoding="utf-8") as f:
        json.dump(history, f, ensure_ascii=False, indent=2)

def get_client_info_from_redis(client_id: str) -> Dict[str, Any]:
    """Get stored client information from Redis"""
    try:
        redis_client: redis.Redis = app.config["SESSION_REDIS"]
        client_info_key = f"client_info:{client_id}"
        existing_info_raw: Optional[bytes] = redis_client.get(client_info_key)
        
        if existing_info_raw:
            return json.loads(existing_info_raw.decode('utf-8'))
        return {}
    except Exception as e:
        logging.error(f"Error getting client info: {e}")
        return {}

def extract_client_name_from_conversation(user_input: str, ai_reply: str) -> Optional[str]:
    """Extract client name from conversation text"""
    combined_text = f"{user_input} {ai_reply}"
    
    # Look for name patterns
    name_patterns = [
        r"(?:my name is|i'm|i am|call me)\s+([A-Za-z][A-Za-z\s]{1,30})",
        r"(?:this is|here is)\s+([A-Za-z][A-ZaZ\s]{1,30})",
    ]
    
    for pattern in name_patterns:
        match = re.search(pattern, combined_text, re.IGNORECASE)
        if match:
            name = match.group(1).strip()
            # Filter out common words that aren't names
            if name.lower() not in ['saba', 'assistant', 'bot', 'ai', 'help', 'meeting']:
                return name
    return None

def is_client_info_complete(client_id: str, email: str, user_input: str, ai_reply: str) -> Dict[str, Any]:
    """
    Check if we have complete client information before scheduling meeting.
    Returns dict with completion status and missing fields.
    """
    client_info = get_client_info_from_redis(client_id)
    
    # Try to extract name from current conversation if not stored
    stored_name = client_info.get('name') or client_info.get('Name')
    extracted_name = extract_client_name_from_conversation(user_input, ai_reply)
    
    client_name = stored_name or extracted_name
    
    # Required fields for meeting scheduling
    required_info = {
        'name': client_name,
        'email': email,
        'meeting_confirmed': False  # Will be set to True after confirmation
    }
    
    missing_fields = []
    
    if not required_info['name']:
        missing_fields.append('name')
    
    # Optional but recommended fields
    business_info = client_info.get('business') or client_info.get('Business') or client_info.get('company')
    phone_info = client_info.get('phone') or client_info.get('Phone') or client_info.get('contact')
    
    return {
        'is_complete': len(missing_fields) == 0,
        'missing_fields': missing_fields,
        'client_info': {
            'name': client_name,
            'email': email,
            'business': business_info,
            'phone': phone_info
        },
        'needs_confirmation': True  # Always require confirmation
    }

def detect_and_schedule_meeting(user_input: str, ai_reply: str, client_id: str) -> Dict[str, Any]:
    """
    Detects email addresses in conversation and schedules Google Meet only after confirming complete client information.
    Returns meeting details if email was detected and meeting scheduling was attempted.
    """
    # 📧 EMAIL PATTERN: Enhanced to catch various email formats
    email_pattern = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
    
    # Check for email in user input
    user_emails = re.findall(email_pattern, user_input)
    
    # 🔍 ENHANCED: Also check conversation history for emails if none found in current input
    conversation_emails = []
    if not user_emails:
        try:
            # Read the conversation file to check for emails in full conversation history
            conversation_files = [f for f in os.listdir('client_data') if f.startswith(f'conversation_{client_id}')]
            if conversation_files:
                latest_conversation = max(conversation_files, key=lambda x: os.path.getctime(os.path.join('client_data', x)))
                conversation_path = os.path.join('client_data', latest_conversation)
                
                with open(conversation_path, 'r', encoding='utf-8') as f:
                    full_conversation = f.read()
                
                conversation_emails = re.findall(email_pattern, full_conversation)
                # Filter out the official email to get client emails only
                conversation_emails = [email for email in conversation_emails if email != "khanjawadkhalid@gmail.com"]
                
                if conversation_emails:
                    logging.info(f"📧 Email found in conversation history: {conversation_emails[0]} for client: {client_id}")
        except Exception as e:
            logging.error(f"❌ Error reading conversation history: {e}")
    
    # Use emails from current input or conversation history
    detected_emails = user_emails or conversation_emails
    
    # Check for meeting-related keywords in the conversation
    meeting_keywords = [
        'meeting', 'schedule', 'google meet', 'meet link', 'appointment', 
        'call', 'discussion', 'consultation', 'demo', 'presentation'
    ]
    
    has_meeting_context = any(keyword in user_input.lower() or keyword in ai_reply.lower() 
                            for keyword in meeting_keywords)
    
    # If email detected and meeting context exists
    if detected_emails and has_meeting_context:
        email = detected_emails[0]  # Use the first email found
        logging.info(f"📧 Email detected: {email} for client: {client_id}")
        
        # Check if we have complete client information first
        info_check = is_client_info_complete(client_id, email, user_input, ai_reply)
        
        if not info_check['is_complete']:
            # Return request for missing information instead of scheduling meeting
            missing = ', '.join(info_check['missing_fields'])
            return {
                "success": False,
                "needs_info": True,
                "missing_fields": info_check['missing_fields'],
                "message": f"I need to collect some information before scheduling your meeting. Please provide: {missing}"
            }
        
        # Check for explicit confirmation keywords
        confirmation_keywords = [
            'yes', 'confirm', 'proceed', 'go ahead', 'schedule it', 'book it', 
            'correct', 'right', 'that\'s right', 'exactly', 'perfect'
        ]
        
        has_confirmation = any(keyword in user_input.lower() for keyword in confirmation_keywords)
        
        # If we don't have explicit confirmation, request it
        if info_check['needs_confirmation'] and not has_confirmation:
            client_info = info_check['client_info']
            return {
                "success": False,
                "needs_confirmation": True,
                "client_info": client_info,
                "message": f"Before I schedule your meeting, let me confirm your details:\n\n" +
                          f"📝 **Name:** {client_info['name']}\n" +
                          f"📧 **Email:** {client_info['email']}\n" +
                          (f"🏢 **Business:** {client_info['business']}\n" if client_info.get('business') else "") +
                          (f"📱 **Phone:** {client_info['phone']}\n" if client_info.get('phone') else "") +
                          f"\nIs this information correct? Please confirm with 'Yes' and I'll schedule your meeting immediately."
            }
        
        # Extract time/date information from the conversation
        meeting_time_info = extract_meeting_time_info(user_input, ai_reply)
        
        if meeting_time_info:
            try:
                # Create Google Meet event
                summary = f"SABA Meeting with {client_id}"
                start_time = meeting_time_info['start_time']
                end_time = meeting_time_info['end_time']
                # Always include the official email along with the client's email
                official_email = "khanjawadkhalid@gmail.com"
                attendees = [email, official_email]
                
                # Use the existing Google Calendar function
                result = create_meet_event(summary, start_time, end_time, attendees)
                
                if result and not result.get("error"):
                    logging.info(f"✅ Meeting scheduled successfully for {email}")
                    logging.info(f"📅 Meeting details: {result}")
                    
                    # Check if there's a note about manual invitation
                    if result.get("note"):
                        logging.warning(f"⚠️ Note: {result['note']}")
                    
                    # Optionally save meeting info to leads
                    save_meeting_to_leads(client_id, email, meeting_time_info, result)
                    return {
                        "success": True,
                        "email": email,
                        "meeting_link": result.get("hangoutLink"),
                        "calendar_link": result.get("htmlLink"),
                        "meeting_details": meeting_time_info,
                        "event_id": result.get("id")
                    }
                else:
                    error_msg = result.get('error', 'Unknown error')
                    logging.error(f"❌ Failed to create meeting: {error_msg}")
                    
                    # If it's a service account permission error, try OAuth directly
                    if "Service accounts cannot invite attendees" in error_msg or "403" in error_msg:
                        logging.info("🔄 Attempting OAuth fallback for attendee invitation...")
                        try:
                            oauth_result = create_meet_event_oauth(summary, start_time, end_time, attendees)
                            if oauth_result and not oauth_result.get("error"):
                                logging.info(f"✅ OAuth Meeting scheduled successfully for {email}")
                                save_meeting_to_leads(client_id, email, meeting_time_info, oauth_result)
                                return {
                                    "success": True,
                                    "email": email,
                                    "meeting_link": oauth_result.get("hangoutLink"),
                                    "calendar_link": oauth_result.get("htmlLink"),
                                    "meeting_details": meeting_time_info,
                                    "event_id": oauth_result.get("id")
                                }
                            else:
                                logging.error(f"❌ OAuth fallback also failed: {oauth_result.get('error', 'Unknown error')}")
                        except Exception as oauth_ex:
                            logging.error(f"❌ OAuth fallback exception: {oauth_ex}")
                    
            except Exception as e:
                logging.error(f"❌ Exception during meeting creation: {e}")
    
    return {"success": False}

def extract_meeting_time_info(user_input: str, ai_reply: str) -> Optional[Dict[str, str]]:
    """
    Extracts meeting time information from the conversation.
    Returns dict with start_time and end_time in ISO format, or None if not found.
    """
    combined_text = f"{user_input} {ai_reply}".lower()
    
    # Look for specific time patterns
    time_patterns = [
        r'(\d{1,2}):(\d{2})\s*(am|pm)',  # 5:00 PM
        r'(\d{1,2})\s*(am|pm)',          # 5 PM
        r'shaam\s*(\d{1,2})',            # Urdu: shaam 5
        r'(\d{1,2})\s*baje',             # Urdu: 5 baje
    ]
    
    # Look for date patterns
    date_patterns = [
        r'(\d{1,2})\s*june',            # "5 June" or "5th June"
        r'june\s*(\d{1,2})',            # "June 5"
        r'(\d{1,2})th\s*june',          # "5th June"
        r'june\s*(\d{1,2})th',          # "June 5th"
        r'today',
        r'tomorrow',
        r'next\s*(week|monday|tuesday|wednesday|thursday|friday)',
    ]
    
    # Current time for calculations
    now = datetime.now()
    default_duration_hours = 1  # Default 1-hour meeting
    
    # Try to find time
    meeting_hour = None
    is_pm = False
    
    for pattern in time_patterns:
        match = re.search(pattern, combined_text)
        if match:
            if len(match.groups()) >= 3:  # Full time with AM/PM
                meeting_hour = int(match.group(1))
                is_pm = match.group(3).lower() == 'pm'
            elif len(match.groups()) >= 2:  # Hour with AM/PM
                meeting_hour = int(match.group(1))
                is_pm = match.group(2).lower() == 'pm'
            elif len(match.groups()) >= 1:  # Just hour (Urdu patterns)
                meeting_hour = int(match.group(1))
                is_pm = True  # Assume PM for Urdu patterns
            break
    
    # Try to find date
    meeting_date = None
    for pattern in date_patterns:
        match = re.search(pattern, combined_text)
        if match:
            if 'june' in pattern:
                day = int(match.group(1)) if match.group(1) else int(match.group(0).replace('june', '').strip())
                meeting_date = datetime(now.year, 6, day)  # June
            elif pattern == 'today':
                meeting_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
            elif pattern == 'tomorrow':
                meeting_date = (now + timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
            break
    
    # Default to tomorrow if no date found
    if not meeting_date:
        meeting_date = (now + timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
    
    # Default to 5 PM if no time found
    if not meeting_hour:
        meeting_hour = 17  # 5 PM
        is_pm = True
    
    # Adjust for AM/PM
    if is_pm and meeting_hour < 12:
        meeting_hour += 12
    elif not is_pm and meeting_hour == 12:
        meeting_hour = 0
    
    # Create start and end times
    start_datetime = meeting_date.replace(hour=meeting_hour, minute=0, second=0, microsecond=0)
    end_datetime = start_datetime + timedelta(hours=default_duration_hours)
    
    # Convert to ISO format with Pakistan timezone
    pak_tz = timezone(timedelta(hours=5))  # PKT is UTC+5
    start_iso = start_datetime.replace(tzinfo=pak_tz).isoformat()
    end_iso = end_datetime.replace(tzinfo=pak_tz).isoformat()
    
    return {
        'start_time': start_iso,
        'end_time': end_iso,
        'meeting_date': meeting_date.strftime('%Y-%m-%d'),
        'meeting_time': f"{meeting_hour:02d}:00"
    }

def save_meeting_to_leads(client_id: str, email: str, meeting_info: Dict[str, str], calendar_result: Dict[str, Any], chat_id: Optional[str] = None, chat_summary: Optional[str] = None, name: Optional[str] = None):
    """
    Saves meeting information to the leads database and updates minimal leads file.
    """
    try:
        leads_data = load_leads_from_file()
        meeting_entry = {
            "id": len(leads_data) + 1,
            "timestamp": datetime.now().isoformat(),
            "name": name or "N/A",
            "contact_method": "auto_scheduled_meeting",
            "contact_details": f"Email: {email}",
            "details": f"Automatically scheduled Google Meet - Date: {meeting_info['meeting_date']} Time: {meeting_info['meeting_time']} PKT",
            "status": "meeting_scheduled",
            "source_client_id": client_id,
            "calendar_event_id": calendar_result.get("id"),
            "meet_link": calendar_result.get("hangoutLink"),
            "calendar_link": calendar_result.get("htmlLink"),
            "source_log_file": f"conversation_{client_id}.txt"
        }
        leads_data.append(meeting_entry)
        save_leads_to_file(leads_data)
        logging.info(f"📝 Meeting entry saved to leads database: {meeting_entry['id']}")
        # --- Update minimal leads file ---
        phone = extract_phone_from_chat_id(chat_id) if chat_id else None
        update_leads_minimal(
            name=name,
            phone=phone,
            email=email,
            chat_summary=chat_summary,
            meet_link=calendar_result.get("hangoutLink"),
            date=meeting_info.get("meeting_date"),
            meeting_time=meeting_info.get("meeting_time")
        )
    except Exception as e:
        logging.error(f"❌ Failed to save meeting to leads: {e}")

LEADS_MINIMAL_FILE = os.path.join(EXCEL_FOLDER, "leads_minimal.json")

def load_leads_minimal() -> List[Dict[str, Any]]:
    if not os.path.exists(LEADS_MINIMAL_FILE):
        return []
    try:
        with open(LEADS_MINIMAL_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return []

def save_leads_minimal(leads: List[Dict[str, Any]]):
    with open(LEADS_MINIMAL_FILE, "w", encoding="utf-8") as f:
        json.dump(leads, f, ensure_ascii=False, indent=2)

def extract_phone_from_chat_id(chat_id: str) -> Optional[str]:
    match = re.match(r"(\d{10,15})@", chat_id)
    if match:
        return match.group(1)
    return None

def update_leads_minimal(name: Optional[str], phone: Optional[str], email: Optional[str], chat_summary: Optional[str], meet_link: Optional[str], date: Optional[str], meeting_time: Optional[str] = None, conversation_file: Optional[str] = None):
    leads = load_leads_minimal()
    # Find by phone
    lead = next((l for l in leads if l.get("phone") == phone), None)
    if lead:
        if name: lead["name"] = name
        if email: lead["email"] = email
        if chat_summary: lead["chat_summary"] = chat_summary
        if meet_link: lead["meet_link"] = meet_link
        if date: lead["date"] = date
        if meeting_time: lead["meeting_time"] = meeting_time
        
        # Add conversation file reference
        if conversation_file:
            conversation_files = lead.get("conversation_files", [])
            if conversation_file not in conversation_files:
                conversation_files.append(conversation_file)
                lead["conversation_files"] = conversation_files
            lead["last_interaction"] = datetime.now().isoformat()
    else:
        new_id = max([l.get("id", 0) for l in leads] + [0]) + 1
        leads.append({
            "id": new_id,
            "name": name or "",
            "phone": phone or "",
            "email": email or "",
            "chat_summary": chat_summary or "",
            "meet_link": meet_link or "",
            "date": date or "",
            "meeting_time": meeting_time or "",
            "conversation_files": [conversation_file] if conversation_file else [],
            "last_interaction": datetime.now().isoformat(),
            "total_messages": 0
        })
    save_leads_minimal(leads)

def get_conversation_history(phone: Optional[str] = None, email: Optional[str] = None, lead_id: Optional[int] = None) -> List[Dict[str, Any]]:
    """
    Get conversation history for a lead - returns lightweight conversation data for Gemini
    """
    leads = load_leads_minimal()
    
    # Find the lead
    lead = None
    if lead_id:
        lead = next((l for l in leads if l.get("id") == lead_id), None)
    elif phone:
        lead = next((l for l in leads if l.get("phone") == phone), None)
    elif email:
        lead = next((l for l in leads if l.get("email") == email), None)
    
    if not lead:
        return []
    
    conversation_files = lead.get("conversation_files", [])
    conversation_history = []
    
    for conv_file in conversation_files:
        try:
            conversation_path = os.path.join(CLIENT_DATA_FOLDER, conv_file)
            if os.path.exists(conversation_path):
                with open(conversation_path, 'r', encoding='utf-8') as f:
                    content = f.read()
                
                # Extract key conversation points (last few exchanges)
                lines = content.split('\n')
                recent_exchanges = []
                
                # Get last 6 exchanges (3 user + 3 model responses)
                user_lines = [line for line in lines if line.startswith('user:')]
                model_lines = [line for line in lines if line.startswith('model:')]
                
                # Take last 3 of each
                recent_user = user_lines[-3:] if len(user_lines) >= 3 else user_lines
                recent_model = model_lines[-3:] if len(model_lines) >= 3 else model_lines
                
                conversation_history.append({
                    "file": conv_file,
                    "last_user_messages": [msg.replace('user:', '').strip() for msg in recent_user],
                    "last_model_responses": [msg.replace('model:', '').strip() for msg in recent_model],
                    "total_user_messages": len(user_lines),
                    "file_timestamp": conv_file.split('_')[-1].replace('.txt', '') if '_' in conv_file else "unknown"
                })
        except Exception:
            continue
    
    return conversation_history

@app.route("/system-instruction", methods=["GET"])
def get_system_instruction_api():
    instruction = get_instruction()
    if instruction is None:
        return jsonify({"error": "Not found"}), 404
    return jsonify({"system_instruction": instruction})

@app.route("/system-instruction", methods=["POST"])
def update_system_instruction_api():
    data = request.get_json()
    new_instruction = data.get("system_instruction", "")
    old_instruction = get_instruction()
    save_instruction(new_instruction)
    # Save to history if changed
    if old_instruction is not None and old_instruction != new_instruction:
        history = get_history()
        history.append({
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "summary": f"Edited instruction. Previous length: {len(old_instruction)}, New length: {len(new_instruction)}"
        })
        save_history(history)
    return jsonify({"message": "System instruction updated"})

@app.route("/system-instruction", methods=["PUT"])
def create_system_instruction_api():
    if os.path.exists(INSTRUCTION_FILE):
        return jsonify({"error": "File already exists"}), 400
    data = request.get_json()
    text = data.get("system_instruction", "Your default instruction here.")
    save_instruction(text)
    save_history([])  # Start with empty history
    return jsonify({"message": "File created"})

@app.route("/system-instruction/history", methods=["GET"])
def get_system_instruction_history_api():
    history = get_history()
    return jsonify({"history": history})

DEFAULT_SYSTEM_INSTRUCTION = """
You are Saba, a clever, slightly witty, and professional AI assistant Saba as a bot for people's business and IMJD Your Digital Media Partner
... (your full instruction here) ...
"""

if not os.path.exists(INSTRUCTION_FILE):
    save_instruction(DEFAULT_SYSTEM_INSTRUCTION)
    save_history([])

@app.route('/gemini/train', methods=['POST'])
@limiter.limit("200000 per day;10000 per hour")
def gemini_train():
    try:
        data = request.get_json()
        logging.debug(f"Received train request: {data}")
        if not data or "content" not in data:
            logging.warning("🚨 'content' key missing: %s", data)
            return jsonify({"error": "Missing 'content' in request body"}), 400

        user_input: str = data.get("content", "").strip()
        if not user_input:
            logging.warning("🚨 Empty 'content': %s", data)
            return jsonify({"error": "Empty 'content' value"}), 400

        client_id: str = data.get("client_id", "default").replace("@", "_")
        phone: Optional[str] = data.get("phone")
        email: Optional[str] = data.get("email")
        logging.debug(f"Client ID: {client_id}, Phone: {phone}, Email: {email}")

        MAX_INPUT_LENGTH = int(os.getenv("MAX_INPUT_LENGTH", 5000))
        if len(user_input) > MAX_INPUT_LENGTH:
            return jsonify({"error": f"Input exceeds {MAX_INPUT_LENGTH} characters"}), 400

        # 📊 GET CLIENT CONTEXT FROM BOTH JSON FILES
        client_context = get_client_context(client_id, phone, email)
        sales_notes = client_context.get("sales_notes", "")
        existing_client_data = client_context.get("client_data")
        
        # 🎯 CLIENT DEDUPLICATION: Check if this is first interaction
        client_result = get_or_create_client_and_conversation(phone, email, client_id)
        is_new_client = client_result["created"]
        client_info = client_result["client"]
        
        logging.debug(f"📋 Sales Notes: {sales_notes[:100]}..." if sales_notes else "📋 No sales notes found")
        logging.debug(f"👤 Client Info: {client_info.get('name')} (ID: {client_info['id']}) - {'NEW' if is_new_client else 'EXISTING'}")

        redis_client: redis.Redis = app.config["SESSION_REDIS"]
        session_key = f"chat_session:{client_id}"

        session_data_raw: Optional[bytes] = redis_client.get(session_key)  # type: ignore
        conversation_history: List[Dict[str, Any]]
        client_name: Optional[str] = None

        is_new_user = False
        if session_data_raw:
            session_data: Dict[str, Any] = json.loads(session_data_raw.decode('utf-8'))
            conversation_history = session_data.get("conversation_history", [])
            client_name = session_data.get("client_name")
        else:
            # New user - initialize session but don't return early
            is_new_user = True
            conversation_history = []
            client_name = None
            logging.debug(f"Initializing new session for {client_id}")

        # 🆕 HANDLE NEW CLIENT: Return greeting with client ID immediately for first message
        if is_new_client and len(conversation_history) == 0:
            greeting = saba_greeting_with_client_id(client_info)
            logging.info(f"🆕 New client detected, sending greeting with ID: {client_info['id']}")
            
            # Add user input and greeting to conversation history
            conversation_history.append({"role": "user", "parts": [{"text": user_input}]})
            conversation_history.append({"role": "model", "parts": [{"text": greeting}]})
            
            # Save session and conversation
            redis_client.setex(session_key, 86400, json.dumps({
                "conversation_history": conversation_history,
                "client_name": client_name,
            }))
            save_conversation(client_id, conversation_history)
            
            return jsonify({"reply": greeting})

        if not client_name:
            logging.debug(f"Checking for name in input: {user_input}")
            name_match = re.search(r"(?:my name is|i'm|i am)\s+([A-Za-z][A-Za-z\s]*)", user_input, re.IGNORECASE)
            if name_match:
                client_name = name_match.group(1)
                logging.debug(f"Client name stored: {client_name}")
            else:
                logging.debug("No name matched in input")

        conversation_history.append({"role": "user", "parts": [{"text": user_input}]})

        MAX_HISTORY_TURNS = int(os.getenv("MAX_HISTORY_TURNS", 10))
        if len(conversation_history) > MAX_HISTORY_TURNS * 2:
            conversation_history = conversation_history[-(MAX_HISTORY_TURNS * 2):]

        redis_client.setex(session_key, 86400, json.dumps({  # type: ignore
            "conversation_history": conversation_history,
            "client_name": client_name,
        }))
        logging.debug(f"Updated session for {client_id}")

        logging.debug(f"Conversation history: {conversation_history}")
        logging.debug(f"📤 Sending to Gemini: {user_input}")

        url = f"https://generativelanguage.googleapis.com/v1beta/models/{GEMINI_MODEL}:generateContent?key={GEMINI_API_KEY}"
        headers = {"Content-Type": "application/json"}
        
        system_instruction_content = get_instruction()
        if system_instruction_content is None:
            system_instruction_content = DEFAULT_SYSTEM_INSTRUCTION

        # 🎯 INJECT CLIENT CONTEXT INTO SYSTEM INSTRUCTION
        context_injection = ""
        
        if sales_notes:
            context_injection += f"\n\n📝 SALES TEAM NOTES FOR THIS CLIENT:\n{sales_notes}\n"
        
        if existing_client_data:
            context_injection += f"\n\n👤 CLIENT INFORMATION:\n"
            context_injection += f"- Name: {existing_client_data.get('name', 'Unknown')}\n"
            context_injection += f"- Phone: {existing_client_data.get('phone', 'N/A')}\n"
            context_injection += f"- Email: {existing_client_data.get('email', 'N/A')}\n"
            context_injection += f"- Previous Summary: {existing_client_data.get('chat_summary', 'N/A')}\n"
            context_injection += f"- Last Interaction: {existing_client_data.get('last_interaction', 'N/A')}\n"
            
            meeting_info = client_context.get("meeting_info", {})
            if meeting_info.get("meet_link"):
                context_injection += f"- Meeting Scheduled: {meeting_info.get('meeting_date')} at {meeting_info.get('meeting_time')}\n"
                context_injection += f"- Meeting Link: {meeting_info.get('meet_link')}\n"
        
        if context_injection:
            context_injection += "\nUse this information to provide personalized, context-aware responses. Reference previous conversations and notes when relevant.\n"
            system_instruction_content += context_injection
            logging.debug(f"💡 Injected client context: {len(context_injection)} characters")

        # 📅 DYNAMIC DATE INJECTION: Update the date context in the system instruction
        current_date = datetime.now().strftime("%B %d, %Y")  # e.g., "June 4, 2025"
        current_time_zone = "Pakistan Standard Time (PST/PKT)"
        
        # Replace the static date in the system instruction with the current date
        date_pattern = r"📅 CURRENT DATE & TIME CONTEXT\s*\nToday's date is: [^\n]+\nCurrent time zone: [^\n]+"
        updated_date_context = f"📅 CURRENT DATE & TIME CONTEXT\nToday's date is: {current_date}\nCurrent time zone: {current_time_zone}"
        
        if re.search(date_pattern, system_instruction_content):
            system_instruction_content = re.sub(date_pattern, updated_date_context, system_instruction_content)
        else:
            # If no date context found, add it at the beginning after the header
            lines = system_instruction_content.split('\n')
            if len(lines) > 8:  # Insert after the header section
                lines.insert(9, f"\n{updated_date_context}")
                system_instruction_content = '\n'.join(lines)
        
        logging.debug(f"📅 Injected current date: {current_date}")

        payload: Dict[str, Any] = {
            "systemInstruction": {"parts": [{"text": system_instruction_content}]},
            "contents": conversation_history
        }

        response = requests.post(url, headers=headers, json=payload)

        if response.status_code == 200:
            res_json = response.json()
            try:
                if not res_json.get("candidates") or not res_json["candidates"][0].get("content") or not res_json["candidates"][0]["content"].get("parts"):
                    logging.error("⚠️ Missing content structure: %s", res_json)
                    finish_reason = res_json.get("candidates", [{}])[0].get("finishReason", "UNKNOWN")
                    if finish_reason != "STOP":
                        block_reason = res_json.get("promptFeedback", {}).get("blockReason", "NONE")
                        logging.warning(f"Prompt blocked: {block_reason}, Details: {res_json.get('promptFeedback')}")
                        return jsonify({
                            "error": "Content blocked due to safety policies.",
                            "details": "Please revise your input and try again."
                        }), 400
                    return jsonify({"error": f"Response issue. Finish Reason: {finish_reason}"}), 500

                reply: str = res_json["candidates"][0]["content"]["parts"][0]["text"]

                # 🔄 UPDATE CLIENT DATA IN LEADS_MINIMAL.JSON
                try:
                    # Extract client information from the full conversation history
                    full_conversation = user_input  # Start with current input
                    try:
                        # Read the conversation file to get full context
                        conversation_files = [f for f in os.listdir('client_data') if f.startswith(f'conversation_{client_id}')]
                        if conversation_files:
                            latest_conversation = max(conversation_files, key=lambda x: os.path.getctime(os.path.join('client_data', x)))
                            conversation_path = os.path.join('client_data', latest_conversation)
                            
                            with open(conversation_path, 'r', encoding='utf-8') as f:
                                full_conversation = f.read()
                                logging.info(f"📖 Using full conversation history for data extraction: {len(full_conversation)} characters")
                    except Exception as e:
                        logging.warning(f"⚠️ Could not read full conversation, using current input only: {e}")
                    
                    extracted_info = extract_client_info_from_conversation(full_conversation, reply)
                    
                    # Update total messages count
                    if existing_client_data:
                        extracted_info["total_messages"] = existing_client_data.get("total_messages", 0) + 1
                    else:
                        extracted_info["total_messages"] = 1
                    
                    # Update the client data
                    success = update_client_data(client_id, extracted_info, phone, email)
                    if success:
                        logging.info(f"✅ Updated client data for {client_id}")
                    else:
                        logging.warning(f"⚠️ Failed to update client data for {client_id}")
                        
                except Exception as e:
                    logging.error(f"❌ Error updating client data: {e}")
                    # Don't break the conversation flow

                if "what is your" in user_input.lower() or "can you tell me your" in user_input.lower():
                    field_match = re.search(r"(what is your|can you tell me your)\s+(.+)", user_input, re.IGNORECASE)
                    if field_match:
                        field = field_match.group(2).strip().capitalize()
                        client_data_item: Dict[str, str] = {field: reply}
                        client_info_key = f"client_info:{client_id}"
                        existing_info_raw: Optional[bytes] = redis_client.get(client_info_key)  # type: ignore
                        existing_info: Dict[str, Any]
                        if existing_info_raw:
                            existing_info = json.loads(existing_info_raw.decode('utf-8'))
                        else:
                            existing_info = {}
                        existing_info.update(client_data_item)
                        redis_client.setex(client_info_key, 86400 * 7, json.dumps(existing_info))  # type: ignore
                        if len(existing_info) >= 3:
                            save_to_excel(existing_info, client_id)

                if "I'm Saba" in reply or "May I have your name" in reply:
                    logging.warning(f"Possible repetition in response: {reply}")

                # 📧 AUTOMATIC EMAIL DETECTION & GOOGLE MEET SCHEDULING WITH CONFIRMATION
                meeting_result = {"success": False}
                try:
                    meeting_result = detect_and_schedule_meeting(user_input, reply, client_id)
                    
                    if meeting_result.get("success"):
                        logging.info(f"📧 Email detected and meeting scheduled for client: {client_id}")
                        
                        # 🔗 ADD MEETING DETAILS TO THE AI RESPONSE
                        meeting_link = meeting_result.get("meeting_link")
                        calendar_link = meeting_result.get("calendar_link")
                        client_email = meeting_result.get("email")
                        meeting_details = meeting_result.get("meeting_details", {})
                        
                        if meeting_link:
                            # Append meeting information to the AI's response
                            meeting_info = f"\n\n✅ **Meeting Scheduled Successfully!**\n\n"
                            meeting_info += f"📧 **Invitation sent to:** {client_email} and khanjawadkhalid@gmail.com\n"
                            meeting_info += f"📅 **Date & Time:** {meeting_details.get('meeting_date', 'TBD')} at {meeting_details.get('meeting_time', 'TBD')} PKT\n"
                            meeting_info += f"🔗 **Google Meet Link:** {meeting_link}\n"
                            if calendar_link:
                                meeting_info += f"📅 **Add to Calendar:** {calendar_link}\n"
                            meeting_info += f"\nPlease save this meeting link and join at the scheduled time. I'll see you there!"
                            
                            # Add the meeting info to the original AI reply
                            reply = reply + meeting_info
                    
                    elif meeting_result.get("needs_info"):
                        logging.info(f"📝 Missing client information for meeting: {client_id}")
                        # Append information request to the AI's response
                        info_request = f"\n\n📝 **Meeting Information Required**\n\n"
                        info_request += meeting_result.get("message", "I need some additional information before scheduling your meeting.")
                        reply = reply + info_request
                    
                    elif meeting_result.get("needs_confirmation"):
                        logging.info(f"📋 Requesting confirmation for meeting: {client_id}")
                        # Append confirmation request to the AI's response
                        confirmation_request = f"\n\n📋 **Please Confirm Your Details**\n\n"
                        confirmation_request += meeting_result.get("message", "Please confirm your information before I schedule the meeting.")
                        reply = reply + confirmation_request
                            
                except Exception as e:
                    logging.error(f"❌ Error in automatic meeting scheduling: {e}")
                    # Don't break the conversation flow if meeting scheduling fails

                conversation_history.append({"role": "model", "parts": [{"text": reply}]})
                redis_client.setex(session_key, 86400, json.dumps({  # type: ignore
                    "conversation_history": conversation_history,
                    "client_name": client_name,
                }))
                save_conversation(client_id, conversation_history)
                logging.debug("Session updated with response")
                
                # Return the natural AI response without forced greetings
                return jsonify({"reply": reply})
            except (KeyError, IndexError) as e:
                logging.error("⚠️ Unexpected response structure: %s, Error: %s", res_json, e)
                return jsonify({"error": "Unexpected response format from Gemini."}), 500
        else:
            logging.error("🔥 Gemini API Error (%s): %s", response.status_code, response.text)
            try:
                return jsonify({"error": "Gemini API request failed.", "details": "Please try again later."}), response.status_code
            except requests.exceptions.JSONDecodeError:
                return jsonify({"error": "Gemini API request failed.", "details": response.text}), response.status_code

    except Exception as e:
        logging.exception(f"🔥 Unhandled exception in /gemini/train: {e}")
        return jsonify({"error": "Internal server error. Please try again."}), 500

@app.route('/gemini/reset', methods=['POST'])
@limiter.limit("20000 per day;5000 per hour")
def reset_conversation():
    try:
        data = request.get_json()
        client_id: str = data.get("client_id", "default").replace("@", "_") if data else "default"
        logging.debug(f"Reset request for client_id: {client_id}")
        redis_client: redis.Redis = app.config["SESSION_REDIS"]
        session_key = f"chat_session:{client_id}"
        redis_client.delete(session_key)  # type: ignore
        logging.debug(f"Deleted session for {client_id}")

        # Clear client info in Redis
        client_info_key = f"client_info:{client_id}"
        redis_client.delete(client_info_key)  # type: ignore

        # Also clear cookie-based session
        session.pop('conversation_history', None)  # type: ignore
        session.pop('client_name', None)  # type: ignore
        session.pop('session_id', None)  # type: ignore
        session.pop('client_info', None)  # type: ignore
        session.modified = True
        logging.debug("Cookie-based session reset")
        return jsonify({"message": "Conversation reset. Please start anew."})
    except Exception as e:
        logging.exception(f"🔥 Reset endpoint error: {e}")
        return jsonify({"error": "Failed to reset conversation."}), 500

@app.route('/gemini/save-info', methods=['POST'])
@limiter.limit("1000 per hour")
def save_client_info():
    try:
        data = request.get_json()
        client_id: str = data.get("client_id", "default").replace("@", "_") if data else "default"
        
        # For Redis-based sessions
        redis_client: redis.Redis = app.config["SESSION_REDIS"]
        client_info_key = f"client_info:{client_id}"
        existing_info_raw: Optional[bytes] = redis_client.get(client_info_key)  # type: ignore
        
        if existing_info_raw:
            existing_info: Dict[str, Any] = json.loads(existing_info_raw.decode('utf-8'))
            if save_to_excel(existing_info, client_id):
                return jsonify({"message": "Client info saved successfully"})
        
        # For cookie-based sessions
        if 'client_info' in session and isinstance(session['client_info'], dict):
            client_info_from_session: Dict[str, Any] = session['client_info']  # type: ignore
            session_id_str: str = session.get('session_id', 'unknown')  # type: ignore
            if save_to_excel(client_info_from_session, session_id_str):
                return jsonify({"message": "Client info saved successfully"})
        
        return jsonify({"error": "No client information to save"}), 404
    except Exception as e:
        logging.exception(f"Error saving client info: {e}")
        return jsonify({"error": str(e)}), 500

@app.route("/logs", methods=["GET"])
def get_logs():
    logs: List[Dict[str, Any]] = []
    for fname in os.listdir(LOGS_FOLDER):
        if fname.endswith(".txt"):
            path = os.path.join(LOGS_FOLDER, fname)
            with open(path, "r", encoding="utf-8") as f:
                content = f.read()
            flagged = "callback" in content.lower() or "google meet" in content.lower()
            logs.append({
                "filename": fname,
                "flagged": flagged,
                "size": os.path.getsize(path),
                "modified": os.path.getmtime(path)
            })
    return jsonify({"logs": logs})

@app.route("/analyze/emotions", methods=["POST"])
def analyze_text_emotions():
    """
    Analyze emotions in the provided text.
    
    Expected JSON body:
    {
        "text": "Text to analyze"
    }
    """
    try:
        data = request.get_json()
        if not data or "text" not in data:
            return jsonify({"error": "Missing 'text' field in request body"}), 400
            
        text = data["text"]
        emotions = analyze_emotion(text)
        
        if "error" in emotions:
            return jsonify({"error": emotions["error"]}), 500
            
        return jsonify({
            "text": text,
            "emotions": emotions
        })
        
    except Exception as e:
        logging.error(f"Error in emotion analysis endpoint: {e}")
        return jsonify({"error": str(e)}), 500

@app.route("/logs/clear", methods=["POST"])
def clear_logs():
    removed: List[str] = []
    for fname in os.listdir(LOGS_FOLDER):
        if fname.endswith(".txt"):
            path = os.path.join(LOGS_FOLDER, fname)
            with open(path, "r", encoding="utf-8") as f:
                content = f.read()
            # Don't delete if flagged for callback or meeting
            if "callback" in content.lower() or "google meet" in content.lower():
                continue
            os.remove(path)
            removed.append(fname)
    return jsonify({"removed": removed, "status": "success"})

@app.route("/logs/<filename>", methods=["GET"])
def get_log_content(filename: str) -> Response:
    """Serve the content of a specific log file."""
    # Basic security: ensure filename is just a filename and doesn't try to traverse directories
    if '..' in filename or '/' in filename:
        return make_response(jsonify({"error": "Invalid filename."}), 400)
    
    file_path = os.path.join(LOGS_FOLDER, filename)
    
    if not os.path.exists(file_path) or not filename.endswith(".txt"):
        return make_response(jsonify({"error": "Log file not found or invalid type."}), 404)
    
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            content = f.read()
        return jsonify({"filename": filename, "content": content})
    except Exception as e_log_read:
        logging.error(f"Error reading log file {filename}: {e_log_read}")
        return make_response(jsonify({"error": "Could not read log file."}), 500)

leads: List[Dict[str, Any]] = load_leads_from_file()  # Load leads at startup

@app.route("/leads", methods=["GET", "POST"])
def get_leads() -> Response:
    """Fetch all leads or filter by source_log_file. Also handles POST for compatibility."""
    # Handle POST requests (might be sent by some frontend automation)
    if request.method == "POST":
        client_ip = request.environ.get('HTTP_X_FORWARDED_FOR', request.environ.get('REMOTE_ADDR', 'unknown'))
        user_agent = request.environ.get('HTTP_USER_AGENT', 'unknown')
        logging.warning(f"Received POST request to /leads endpoint from {client_ip}, User-Agent: {user_agent}, treating as GET")
        
        # Check if there's any POST data that we should handle
        if request.is_json:
            post_data = request.get_json()
            logging.info(f"POST data received: {post_data}")
    
    source_log_file = request.args.get('source_log_file')
    if source_log_file:
        # Ensure the filename is safe, similar to get_log_content
        if '..' in source_log_file or '/' in source_log_file:
            return make_response(jsonify({"error": "Invalid source_log_file parameter."}), 400)
        
        filtered_leads = [lead for lead in leads if lead.get("source_log_file") == source_log_file]
        if not filtered_leads: # If no leads match, return empty list, not 404, as it's a valid query
            return jsonify({"leads": []})
        return jsonify({"leads": filtered_leads})
    
    return jsonify({"leads": leads})

@app.route("/leads-minimal", methods=["GET"])
def get_leads_minimal_api():
    """Return minimal leads data for frontend display"""
    leads = load_leads_minimal()
    return jsonify({"leads": leads})

@app.route("/conversation-history/<int:lead_id>", methods=["GET"])
def get_conversation_history_api(lead_id: int):
    """Get conversation history for a specific lead - optimized for Gemini API"""
    history = get_conversation_history(lead_id=lead_id)
    return jsonify({"conversation_history": history})

@app.route("/lead-context", methods=["POST"])
def get_lead_context_api():
    """
    Get lead context by phone/email - for Gemini to quickly lookup previous conversations
    Returns lightweight lead data + recent conversation points
    """
    data = request.get_json()
    phone = data.get("phone")
    email = data.get("email")
    
    if not phone and not email:
        return jsonify({"error": "Phone or email required"}), 400
    
    # Get lead data
    leads = load_leads_minimal()
    
    # For demo purposes, also check test data
    try:
        test_leads_path = os.path.join(CLIENT_DATA_FOLDER, "leads_minimal_test.json")
        if os.path.exists(test_leads_path):
            with open(test_leads_path, 'r', encoding='utf-8') as f:
                test_leads = json.load(f)
                leads.extend(test_leads)
    except Exception:
        pass
    
    lead = None
    if phone:
        lead = next((l for l in leads if l.get("phone") == phone), None)
    if not lead and email:
        lead = next((l for l in leads if l.get("email") == email), None)
    
    if not lead:
        return jsonify({"lead_found": False, "message": "No previous conversation found"})
    
    # Get recent conversation history
    conversation_history = get_conversation_history(phone=phone, email=email)
    
    # Load agent notes for this lead
    notes_file = os.path.join(CLIENT_DATA_FOLDER, "lead_notes.json")
    agent_notes = ""
    if os.path.exists(notes_file):
        try:
            with open(notes_file, 'r', encoding='utf-8') as f:
                all_notes = json.load(f)
                agent_notes = all_notes.get(str(lead.get("id")), "")
        except Exception:
            agent_notes = ""
    
    # Return lightweight context for Gemini
    context = {
        "lead_found": True,
        "lead": {
            "id": lead.get("id"),
            "name": lead.get("name"),
            "phone": lead.get("phone"),
            "email": lead.get("email"),
            "chat_summary": lead.get("chat_summary"),
            "last_interaction": lead.get("last_interaction"),
            "total_messages": lead.get("total_messages", 0),
            "has_meeting": bool(lead.get("meet_link")),
            "agent_notes": agent_notes  # Include agent notes in context
        },
        "recent_conversations": conversation_history
    }
    
    return jsonify(context)

@app.route("/link-conversation", methods=["POST"])
def link_conversation_api():
    """
    Link a conversation file to a lead - to be called when new conversations are created
    """
    data = request.get_json()
    conversation_file = data.get("conversation_file")
    phone = data.get("phone")
    email = data.get("email")
    name = data.get("name")
    
    if not conversation_file:
        return jsonify({"error": "conversation_file required"}), 400
    
    link_conversation_to_lead(conversation_file, phone, email, name)
    return jsonify({"success": True, "message": "Conversation linked to lead"})

@app.route("/saba/book-meeting", methods=["POST"])
def saba_book_meeting():
    data = request.json
    if not data:
        return jsonify({"error": "Missing request data."}), 400
    summary = data.get("title", "Meeting with SABA")
    start_iso = data.get("start")
    end_iso = data.get("end")
    attendees = data.get("attendees", [])
    # Validate required fields
    if not start_iso or not end_iso:
        return jsonify({"error": "Missing start or end time."}), 400
    if not isinstance(attendees, list) or not all(isinstance(a, str) for a in attendees):
        return jsonify({"error": "Attendees must be a list of emails."}), 400
    # Validate ISO format and ensure times are in the future
    try:
        start_dt = datetime.fromisoformat(start_iso)
        end_dt = datetime.fromisoformat(end_iso)
        if start_dt < datetime.now():
            return jsonify({"error": "Start time must be in the future."}), 400
        if end_dt <= start_dt:
            return jsonify({"error": "End time must be after start time."}), 400
        result = create_meet_event(summary, start_iso, end_iso, attendees)
        if result and not result.get("error"):
            return jsonify({"message": "Meeting scheduled successfully.", "event": result})
        else:
            # Fallback: try OAuth if service account fails to invite attendees
            if "Service accounts cannot invite attendees" in result.get("error", "") or "403" in result.get("error", ""):
                try:
                    oauth_result = create_meet_event_oauth(summary, start_iso, end_iso, attendees)
                    if oauth_result and not oauth_result.get("error"):
                        return jsonify({"message": "Meeting scheduled successfully (OAuth fallback).", "event": oauth_result})
                    else:
                        return jsonify({"error": oauth_result.get("error", "Unknown error (OAuth fallback)")}), 500
                except Exception as oauth_ex:
                    return jsonify({"error": f"OAuth fallback failed: {oauth_ex}"}), 500
            return jsonify({"error": result.get("error", "Unknown error")}), 500
    except Exception as general_error:
        return jsonify({"error": f"Unexpected error: {general_error}"}), 500

@app.route("/saba/book-meeting-oauth", methods=["POST"])
def saba_book_meeting_oauth():
    data = request.json
    if not data:
        return jsonify({"error": "Missing request data."}), 400
    summary = data.get("title", "Meeting with SABA")
    start_iso = data.get("start")
    end_iso = data.get("end")
    attendees = data.get("attendees", [])
    if not start_iso or not end_iso:
        return jsonify({"error": "Missing start or end time."}), 400
    if not isinstance(attendees, list) or not all(isinstance(a, str) for a in attendees):
        return jsonify({"error": "Attendees must be a list of emails."}), 400
    try:
        start_dt = datetime.fromisoformat(start_iso)
        end_dt = datetime.fromisoformat(end_iso)
        if start_dt < datetime.now():
            return jsonify({"error": "Start time must be in the future."}), 400
        if end_dt <= start_dt:
            return jsonify({"error": "End time must be after start time."}), 400
        oauth_result = create_meet_event_oauth(summary, start_iso, end_iso, attendees)
        if oauth_result and not oauth_result.get("error"):
            return jsonify({"message": "Meeting scheduled successfully (OAuth).", "event": oauth_result})
        else:
            return jsonify({"error": oauth_result.get("error", "Unknown error (OAuth)")}), 500
    except Exception as e:
        return jsonify({"error": f"Unexpected error: {e}"}), 500

@app.route('/authorize')
def authorize():
    # Construct the redirect_uri using the current request's root URL
    redirect_uri_val = f"{request.url_root.rstrip('/')}" + url_for('oauth2callback')
    flow = Flow.from_client_secrets_file(
        CLIENT_SECRETS_FILE,
        scopes=SCOPES,
        redirect_uri=redirect_uri_val
    )
    authorization_url, state = flow.authorization_url(
        access_type='offline',
        include_granted_scopes='true',
        prompt='consent'
    )
    session['state'] = state
    return redirect(authorization_url)

@app.route('/oauth2callback')
def oauth2callback():
    state = session.get('state')
    redirect_uri_val = f"{request.url_root.rstrip('/')}" + url_for('oauth2callback')
    flow = Flow.from_client_secrets_file(
        CLIENT_SECRETS_FILE,
        scopes=SCOPES,
        state=state,
        redirect_uri=redirect_uri_val
    )
    flow.fetch_token(authorization_response=request.url)
    credentials = flow.credentials
    # Save the credentials for future use
    with open('token.pickle', 'wb') as token:
        pickle.dump(credentials, token)
    return redirect(url_for('leads_management'))

def create_meet_event_oauth(summary: str, start_time: str, end_time: str, attendees: List[str]) -> Dict[str, Any]:
    with open('token.pickle', 'rb') as token:
        credentials = pickle.load(token)
    service = build('calendar', 'v3', credentials=credentials)
    event = {
        'summary': summary,
        'start': {'dateTime': start_time, 'timeZone': 'Asia/Karachi'},
        'end': {'dateTime': end_time, 'timeZone': 'Asia/Karachi'},
        'attendees': [{'email': a} for a in attendees],
        'conferenceData': {
            'createRequest': {
                'requestId': str(uuid.uuid4()),
                'conferenceSolutionKey': {'type': 'hangoutsMeet'}
            }
        }
    }
    created_event = service.events().insert(
        calendarId='primary',
        body=event,
        conferenceDataVersion=1
    ).execute()
    return created_event

@app.route("/leads/reports/generate", methods=["POST"])
def generate_daily_lead_report() -> Response:
    try:
        leads_data: List[Dict[str, Any]] = load_leads_from_file()
        if not leads_data:
            return make_response(jsonify({"message": "No leads data available to generate a report."}), 200)

        now = datetime.now()
        twenty_four_hours_ago = now - timedelta(hours=24)
        
        recent_leads: List[Dict[str, Any]] = []
        for lead_item in leads_data:  # Renamed lead to lead_item to avoid conflict
            lead_timestamp_str = lead_item.get("timestamp")
            if lead_timestamp_str:
                try:
                    # Attempt to parse ISO format, handling potential 'Z' and microsecond variations
                    if '.' in lead_timestamp_str and ('Z' in lead_timestamp_str.upper() or '+' in lead_timestamp_str or '-' in lead_timestamp_str[10:]):
                        lead_datetime = datetime.fromisoformat(lead_timestamp_str.replace('Z', '+00:00'))
                        # If timezone aware, convert to naive UTC then to system's local naive time for comparison
                        if lead_datetime.tzinfo:
                            lead_datetime = lead_datetime.astimezone(None).replace(tzinfo=None)
                    elif '.' in lead_timestamp_str:  # Handle cases with microseconds but no explicit timezone
                        try:
                            lead_datetime = datetime.fromisoformat(lead_timestamp_str)
                        except ValueError:  # Try parsing without microseconds if full parse fails
                            lead_datetime = datetime.fromisoformat(lead_timestamp_str.split('.')[0])
                    else:  # Handle cases without microseconds
                        lead_datetime = datetime.fromisoformat(lead_timestamp_str)

                    if lead_datetime >= twenty_four_hours_ago:
                        recent_leads.append(lead_item)
                except ValueError as e_parse:
                    logging.warning(f"Could not parse timestamp '{lead_timestamp_str}' for lead ID {lead_item.get("id")}: {e_parse}")
                    continue  # Skip this lead if timestamp is unparseable

        if not recent_leads:
            return make_response(jsonify({"message": "No leads found in the last 24 hours."}), 200)

        wb = Workbook()
        ws_candidate = wb.active
        if ws_candidate is None:
            ws: Worksheet = wb.create_sheet()
        else:
            ws: Worksheet = ws_candidate
        ws.title = "Daily Leads Report"

        headers = ["ID", "Timestamp", "Name", "Contact Method", "Contact Details", "Details", "Status", "Source Client ID", "Source Log File"]
        ws.append(headers)
        # Enhanced header formatting
        for cell in ws[1]:  # type: ignore 
            if cell:
                cell.font = Font(bold=True, color="FFFFFF")
                try:
                    from openpyxl.styles import PatternFill
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                except ImportError:
                    pass  # Skip if PatternFill not available

        # Add data with better formatting
        for lead_entry in recent_leads:
            details = lead_entry.get("details", "")
            # Ensure full details are preserved (no truncation)
            row_data = [
                lead_entry.get("id"),
                lead_entry.get("timestamp"),
                lead_entry.get("name") or "N/A",
                lead_entry.get("contact_method"),
                lead_entry.get("contact_details") or "",
                details,  # Full details without truncation
                lead_entry.get("status"),
                lead_entry.get("source_client_id"),
                lead_entry.get("source_log_file")
            ]
            ws.append(row_data)
        
        # Auto-adjust column widths for better readability
        try:
            from openpyxl.utils import get_column_letter
            for col in range(1, len(headers) + 1):
                column_letter = get_column_letter(col)
                max_length = len(headers[col-1]) if col <= len(headers) else 10
                
                for row in range(2, len(recent_leads) + 2):
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value:
                        max_length = max(max_length, min(len(str(cell_value)), 100))
                
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        except ImportError:
            pass  # Skip column width adjustment if not available
        
        report_filename = f"daily_leads_{now.strftime('%Y-%m-%d_%H%M%S')}.xlsx"
        report_filepath = os.path.join(DAILY_REPORTS_FOLDER, report_filename)
        wb.save(report_filepath)
        
        logging.info(f"Daily lead report generated: {report_filepath}")
        return make_response(jsonify({"message": "Daily lead report generated successfully.", "filename": report_filename}), 201)

    except Exception as e_report:
        logging.exception("Error generating daily lead report")
        return make_response(jsonify({"error": f"Failed to generate report: {str(e_report)}"}), 500)

@app.route("/leads/reports", methods=["GET"])
def list_lead_reports() -> Response:
    try:
        if not os.path.exists(DAILY_REPORTS_FOLDER):
            os.makedirs(DAILY_REPORTS_FOLDER)  # Ensure directory exists
            return jsonify({"reports": []})

        reports = [f for f in os.listdir(DAILY_REPORTS_FOLDER) if f.endswith(".xlsx") and os.path.isfile(os.path.join(DAILY_REPORTS_FOLDER, f))]
        
        # Sort reports by modification time, newest first
        reports_with_details: List[Dict[str, Any]] = []  # Explicitly type reports_with_details
        for report_name in reports:
            try:
                filepath = os.path.join(DAILY_REPORTS_FOLDER, report_name)
                reports_with_details.append({
                    "filename": report_name,
                    "size": os.path.getsize(filepath),
                    "modified_timestamp": os.path.getmtime(filepath)
                })
            except Exception as e_file_stat:
                logging.warning(f"Could not get details for report file {report_name}: {e_file_stat}")
        
        # Sort by 'modified_timestamp' descending
        reports_with_details.sort(key=lambda x: x["modified_timestamp"], reverse=True)
        
        return jsonify({"reports": reports_with_details})
    except Exception as e_list_reports:
        logging.exception("Error listing lead reports")
        return make_response(jsonify({"error": f"Failed to list reports: {str(e_list_reports)}"}), 500)

@app.route("/leads/reports/download/<path:filename>", methods=["GET"])
def download_lead_report(filename: str) -> Response:
    try:
        # Security: Ensure filename is not attempting to traverse directories
        if ".." in filename or filename.startswith("/"):
            return make_response(jsonify({"error": "Invalid filename."}), 400)

        # Check if the DAILY_REPORTS_FOLDER exists
        if not os.path.exists(DAILY_REPORTS_FOLDER) or not os.path.isdir(DAILY_REPORTS_FOLDER):
            logging.error(f"Daily reports folder not found: {DAILY_REPORTS_FOLDER}")
            return make_response(jsonify({"error": "Report folder not found."}), 404)

        # Check if the specific file exists within the directory
        file_path = os.path.join(DAILY_REPORTS_FOLDER, filename)
        if not os.path.isfile(file_path):
            logging.warning(f"Report file not found: {file_path}")
            return make_response(jsonify({"error": "Report file not found."}), 404)
            
        return send_from_directory(DAILY_REPORTS_FOLDER, filename, as_attachment=True)
    except Exception as e_download:
        logging.exception(f"Error downloading report file {filename}")
        return make_response(jsonify({"error": f"Failed to download report: {str(e_download)}"}), 500)

@app.route("/")
def home():
    return "✅ IMJD Gemini API is running."

def generate_chat_summary(conversation_history: List[Dict[str, Any]]) -> str:
    # Simple summary: concatenate first 2-3 user/model messages
    summary_parts = []
    for msg in conversation_history:
        if msg.get("role") in ("user", "model"):
            text = msg.get("parts", [{}])[0].get("text", "")
            if text:
                summary_parts.append(text)
        if len(summary_parts) >= 4:
            break
    return " | ".join(summary_parts)

def link_conversation_to_lead(conversation_file: str, phone: Optional[str] = None, email: Optional[str] = None, name: Optional[str] = None):
    """
    Link a conversation file to a lead in leads_minimal.json
    This function should be called whenever a new conversation is created
    """
    if not phone and not email:
        return  # Need at least one identifier
    
    leads = load_leads_minimal()
    
    # Find existing lead by phone or email
    lead = None
    if phone:
        lead = next((l for l in leads if l.get("phone") == phone), None)
    if not lead and email:
        lead = next((l for l in leads if l.get("email") == email), None)
    
    if lead:
        # Update existing lead
        conversation_files = lead.get("conversation_files", [])
        if conversation_file not in conversation_files:
            conversation_files.append(conversation_file)
            lead["conversation_files"] = conversation_files
        lead["last_interaction"] = datetime.now().isoformat()
        
        # Update message count (count lines starting with 'user:' in the conversation file)
        try:
            conversation_path = os.path.join(CLIENT_DATA_FOLDER, conversation_file)
            if os.path.exists(conversation_path):
                with open(conversation_path, 'r', encoding='utf-8') as f:
                    content = f.read()
                user_messages = len([line for line in content.split('\n') if line.startswith('user:')])
                lead["total_messages"] = lead.get("total_messages", 0) + user_messages
        except Exception:
            pass
    else:
        # Create new lead entry
        new_id = max([l.get("id", 0) for l in leads] + [0]) + 1
        
        # Count messages in conversation file
        total_messages = 0
        try:
            conversation_path = os.path.join(CLIENT_DATA_FOLDER, conversation_file)
            if os.path.exists(conversation_path):
                with open(conversation_path, 'r', encoding='utf-8') as f:
                    content = f.read()
                total_messages = len([line for line in content.split('\n') if line.startswith('user:')])
        except Exception:
            pass
        
        leads.append({
            "id": new_id,
            "name": name or "",
            "phone": phone or "",
            "email": email or "",
            "chat_summary": "Conversation in progress",
            "meet_link": "",
            "date": "",
            "meeting_time": "",
            "meeting_timezone": "",
            "conversation_files": [conversation_file],
            "last_interaction": datetime.now().isoformat(),
            "total_messages": total_messages
        })
    
    save_leads_minimal(leads)

# Client data file paths
LEAD_NOTES_FILE = os.path.join(EXCEL_FOLDER, "lead_notes.json")
LEADS_MINIMAL_FILE = os.path.join(EXCEL_FOLDER, "leads_minimal.json")

def load_lead_notes() -> Dict[str, str]:
    """Load sales team notes from lead_notes.json (read-only for Gemini)"""
    try:
        if os.path.exists(LEAD_NOTES_FILE):
            with open(LEAD_NOTES_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        return {}
    except Exception as e:
        logging.error(f"Error loading lead notes: {e}")
        return {}

def load_leads_minimal() -> List[Dict[str, Any]]:
    """Load main leads data from leads_minimal.json (Gemini can read/write)"""
    try:
        if os.path.exists(LEADS_MINIMAL_FILE):
            with open(LEADS_MINIMAL_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        return []
    except Exception as e:
        logging.error(f"Error loading leads minimal: {e}")
        return []

def save_leads_minimal(leads_data: List[Dict[str, Any]]) -> bool:
    """Save leads data to leads_minimal.json"""
    try:
        with open(LEADS_MINIMAL_FILE, "w", encoding="utf-8") as f:
            json.dump(leads_data, f, indent=2)
        logging.info(f"✅ Leads minimal data saved to {LEADS_MINIMAL_FILE}")
        return True
    except Exception as e:
        logging.error(f"❌ Error saving leads minimal: {e}")
        return False

def find_client_by_contact(phone: str = None, email: str = None, client_id: str = None) -> Optional[Dict[str, Any]]:
    """Find client in leads_minimal.json by phone, email, or conversation file"""
    leads = load_leads_minimal()
    
    for lead in leads:
        # Check phone match (normalize phone numbers)
        if phone and lead.get("phone"):
            # Remove non-digits for comparison
            lead_phone = re.sub(r'[^\d]', '', lead["phone"])
            input_phone = re.sub(r'[^\d]', '', phone)
            if lead_phone == input_phone or lead_phone.endswith(input_phone[-10:]):
                return lead
        
        # Check email match
        if email and lead.get("email") and lead["email"].lower() == email.lower():
            return lead
    
    # NEW: Check conversation files if client_id provided and no match found
    if client_id:
        # First try exact conversation file name match
        for lead in leads:
            conversation_files = lead.get("conversation_files", [])
            for conv_file in conversation_files:
                if client_id in conv_file:
                    logging.info(f"✅ Found client by conversation file: {conv_file}")
                    return lead
        
        # If still no match, check if conversation file exists in filesystem
        # Extract normalized client ID for filesystem search
        normalized_client_id = client_id.replace("@", "_")
        
        if os.path.exists(CLIENT_DATA_FOLDER):
            for filename in os.listdir(CLIENT_DATA_FOLDER):
                if filename.startswith("conversation_") and normalized_client_id in filename:
                    logging.info(f"📁 Found orphaned conversation file: {filename}")
                    # Check if we can extract client info from the conversation file
                    try:
                        with open(os.path.join(CLIENT_DATA_FOLDER, filename), 'r', encoding='utf-8') as f:
                            conversation_content = f.read()
                        
                        # Extract email from conversation file
                        email_pattern = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
                        emails = re.findall(email_pattern, conversation_content)
                        client_emails = [email for email in emails if email.lower() != "khanjawadkhalid@gmail.com"]
                        
                        if client_emails:
                            # Look for existing client with this email
                            for lead in leads:
                                if lead.get("email", "").lower() == client_emails[0].lower():
                                    logging.info(f"🔗 Linked orphaned conversation to existing client via email: {client_emails[0]}")
                                    # Add the conversation file to this client's files
                                    if filename not in lead.get("conversation_files", []):
                                        lead.setdefault("conversation_files", []).append(filename)
                                        save_leads_minimal(leads)
                                    return lead
                    except Exception as e:
                        logging.error(f"❌ Error reading conversation file {filename}: {e}")
    
    return None

def get_client_context(client_id: str, phone: str = None, email: str = None) -> Dict[str, Any]:
    """Get comprehensive client context from both JSON files"""
    context = {
        "sales_notes": "",
        "client_data": None,
        "conversation_summary": "",
        "last_interaction": "",
        "meeting_info": {}
    }
    
    # 1. Get sales team notes (read-only)
    lead_notes = load_lead_notes()
    context["sales_notes"] = lead_notes.get(client_id, "")
    
    # 2. Find client in leads_minimal by phone/email or client_id
    client_data = None
    if phone or email:
        client_data = find_client_by_contact(phone, email)
    
    if not client_data:
        # Try to find by ID
        leads = load_leads_minimal()
        for lead in leads:
            if str(lead.get("id")) == str(client_id):
                client_data = lead
                break
    
    if client_data:
        context["client_data"] = client_data
        context["conversation_summary"] = client_data.get("chat_summary", "")
        context["last_interaction"] = client_data.get("last_interaction", "")
        context["meeting_info"] = {
            "meet_link": client_data.get("meet_link"),
            "meeting_date": client_data.get("meeting_date"),
            "meeting_time": client_data.get("meeting_time"),
            "meeting_timezone": client_data.get("meeting_timezone")
        }
    context["client_id"] = client_id  # Always include client_id in context
    return context

def update_client_data(client_id: str, updates: Dict[str, Any], phone: str = None, email: str = None) -> bool:
    """Update or create client data in leads_minimal.json"""
    try:
        leads = load_leads_minimal()
        
        # Find existing client using enhanced lookup
        client_data = find_client_by_contact(phone, email, client_id)
        client_index = -1
        
        if client_data:
            # Find the index of the existing client
            for i, lead in enumerate(leads):
                if lead.get("id") == client_data.get("id"):
                    client_index = i
                    break
            logging.info(f"✅ Found existing client: {client_data.get('name', 'Unknown')} (ID: {client_data.get('id')})")
        else:
            # Create new client if not found
            max_id = max([lead.get("id", 0) for lead in leads] + [0])
            client_data = {
                "id": max_id + 1,
                "name": "",
                "phone": phone or "",
                "email": email or "",
                "chat_summary": "",
                "meet_link": "",
                "meeting_date": "",
                "meeting_time": "",
                "meeting_timezone": "",
                "conversation_files": [],
                "last_interaction": "",
                "total_messages": 0
            }
            leads.append(client_data)
            client_index = len(leads) - 1
            logging.info(f"🆕 Created new client: ID {client_data['id']}")
        
        # Update client data
        for key, value in updates.items():
            if key in client_data:
                client_data[key] = value
        
        # Update last interaction timestamp
        client_data["last_interaction"] = datetime.now().isoformat()
        
        # Update in the list
        leads[client_index] = client_data
        
        # Save back to file
        return save_leads_minimal(leads)
        
    except Exception as e:
        logging.error(f"❌ Error updating client data: {e}")
        return False

def extract_client_info_from_conversation(conversation: str, reply: str) -> Dict[str, Any]:
    """Extract client information from conversation using AI analysis"""
    extracted_info = {}
    
    # Extract name - prioritize explicit name mentions
    # High priority patterns - explicit name introductions
    priority_name_patterns = [
        r"(?:my name is|i'm|i am|this is|call me)\s+([a-zA-Z]+)",  # Only capture first word to avoid newlines
        r"user:\s+My name is\s+([A-Z][a-zA-Z]+)",  # Specific for Salman's case - only first word
        r"(?:Hi|Hello|Hey),?\s+I'm\s+([A-Z][a-zA-Z]+)",  # Only first word
        r"(?:I'm|This is)\s+([A-Z][a-zA-Z]+)",  # Only first word to avoid capturing too much
    ]
    
    # Lower priority patterns - more general
    general_name_patterns = [
        r"(?:^|\n)([A-Z][a-zA-Z]+)\s+(?:here|speaking)",  # Names followed by "here" or "speaking"
        r"user:\s+([A-Z][a-zA-Z]+)(?:\s|,|$)",  # Names at start of user messages - moved to lower priority
    ]
    
    # Extended false positive filters
    false_positives = [
        'saba', 'ai', 'bot', 'hello', 'help', 'please', 'thank', 'yes', 'no', 
        'ok', 'great', 'good', 'sure', 'model', 'user', 'my', 'name', 'is',
        'relevant', 'product', 'information', 'catalog', 'social', 'media',
        'marketing', 'manager', 'automation', 'business', 'service', 'agency'
    ]
    
    # First try high priority patterns
    for pattern in priority_name_patterns:
        matches = re.findall(pattern, conversation, re.IGNORECASE | re.MULTILINE)
        for match in matches:
            if isinstance(match, tuple):
                name = match[0].strip().title()
            else:
                name = match.strip().title()
            
            # Filter out false positives
            if (len(name) > 1 and len(name) < 30 and 
                not any(word in name.lower().split() for word in false_positives) and 
                not name.isdigit() and
                not name.lower() in ['the', 'and', 'or', 'but', 'with', 'from']):
                extracted_info["name"] = name
                break
        if "name" in extracted_info:
            break
    
    # If no name found with priority patterns, try general patterns
    if "name" not in extracted_info:
        for pattern in general_name_patterns:
            matches = re.findall(pattern, conversation, re.IGNORECASE | re.MULTILINE)
            for match in matches:
                if isinstance(match, tuple):
                    name = match[0].strip().title()
                else:
                    name = match.strip().title()
                
                # More strict filtering for general patterns
                if (len(name) > 1 and len(name) < 30 and 
                    not any(word in name.lower().split() for word in false_positives) and 
                    not name.isdigit() and
                    not name.lower() in ['the', 'and', 'or', 'but', 'with', 'from'] and
                    name.isalpha()):  # Only alphabetic names for general patterns
                    extracted_info["name"] = name
                    break
            if "name" in extracted_info:
                break
    
    # Extract phone
    phone_pattern = r'(\+?[\d\s\-\(\)]{10,15})'
    phone_match = re.search(phone_pattern, conversation)
    if phone_match:
        extracted_info["phone"] = phone_match.group(1).strip()
    
    # Extract email - prioritize client emails over official email
    email_pattern = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
    email_matches = re.findall(email_pattern, conversation)
    if email_matches:
        # Filter out the official email and get client emails
        official_email = "khanjawadkhalid@gmail.com"
        client_emails = [email for email in email_matches if email.lower() != official_email.lower()]
        if client_emails:
            extracted_info["email"] = client_emails[0].lower()  # Use first client email
        elif email_matches:
            extracted_info["email"] = email_matches[0].lower()  # Fallback to any email
    
    # Generate conversation summary
    if len(conversation) > 50:
        # Simple summary - you could enhance this with AI
        summary_parts = []
        if "service" in conversation.lower() or "product" in conversation.lower():
            summary_parts.append("Interested in services")
        if "price" in conversation.lower() or "cost" in conversation.lower():
            summary_parts.append("Asked about pricing")
        if "meeting" in conversation.lower() or "call" in conversation.lower():
            summary_parts.append("Requested meeting")
        
        extracted_info["chat_summary"] = ", ".join(summary_parts) if summary_parts else "General inquiry"
    
    return extracted_info

@app.route("/notes/<client_id>", methods=["POST"])
@limiter.limit("1000 per hour")
def add_client_notes(client_id: str):
    """Add or update sales team notes for a client (lead_notes.json)"""
    try:
        data = request.get_json()
        notes = data.get("notes", "").strip()
        
        if not notes:
            return jsonify({"error": "Notes content required"}), 400
        
        # Load existing notes
        lead_notes = load_lead_notes()
        
        # Update notes for this client
        lead_notes[client_id] = notes
        
        # Save back to file
        try:
            with open(LEAD_NOTES_FILE, "w", encoding="utf-8") as f:
                json.dump(lead_notes, f, indent=2)
            
            logging.info(f"✅ Added notes for client {client_id}: {len(notes)} characters")
            return jsonify({
                "message": "Notes saved successfully",
                "client_id": client_id,
                "notes_length": len(notes)
            })
            
        except Exception as e:
            logging.error(f"❌ Error saving notes: {e}")
            return jsonify({"error": "Failed to save notes"}), 500
            
    except Exception as e:
        logging.error(f"❌ Error in add_client_notes: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route("/notes/<client_id>", methods=["GET"])
def get_client_notes(client_id: str):
    """Get sales team notes for a specific client"""
    try:
        lead_notes = load_lead_notes()
        notes = lead_notes.get(client_id, "")
        
        return jsonify({
            "client_id": client_id,
            "notes": notes,
            "has_notes": bool(notes)
        })
        
    except Exception as e:
        logging.error(f"❌ Error getting notes: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route("/leads/minimal", methods=["GET"])
def get_leads_minimal_for_frontend():
    """Frontend-compatible endpoint for leads_minimal.json"""
    try:
        leads = load_leads_minimal()
        return jsonify({
            "leads": leads,
            "total_count": len(leads)
        })
        
    except Exception as e:
        logging.error(f"❌ Error getting leads for frontend: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route("/notes/sync", methods=["POST"])
def sync_notes_from_frontend():
    """Sync notes from frontend to backend (lead_notes.json)"""
    try:
        data = request.get_json()
        frontend_notes = data.get("notes", {})
        
        if not frontend_notes:
            return jsonify({"message": "No notes to sync"})
        
        # Load existing notes
        lead_notes = load_lead_notes()
        
        # Update with frontend notes
        updated_count = 0
        for client_id, note_content in frontend_notes.items():
            if note_content.strip():  # Only update non-empty notes
                lead_notes[str(client_id)] = note_content.strip()
                updated_count += 1
        
        # Save back to file
        try:
            with open(LEAD_NOTES_FILE, "w", encoding="utf-8") as f:
                json.dump(lead_notes, f, indent=2)
            
            logging.info(f"✅ Synced {updated_count} notes from frontend")
            return jsonify({
                "message": "Notes synced successfully",
                "updated_count": updated_count
            })
            
        except Exception as e:
            logging.error(f"❌ Error saving synced notes: {e}")
            return jsonify({"error": "Failed to save notes"}), 500
            
    except Exception as e:
        logging.error(f"❌ Error in sync_notes_from_frontend: {e}")
        return jsonify({"error": "Internal server error"}), 500

def find_existing_conversation_file(client_id: str) -> Optional[str]:
    """Find the most recent conversation file for a client"""
    try:
        # Check if client exists in leads_minimal and has conversation files
        leads = load_leads_minimal()
        for lead in leads:
            # Extract phone from client_id for comparison
            client_phone = extract_phone_from_chat_id(client_id)
            lead_phone = lead.get("phone", "")
            
            # Compare by phone number or direct ID match
            if (client_phone and lead_phone and 
                client_phone.replace("+", "").replace(" ", "") in lead_phone.replace("+", "").replace(" ", "")):
                conversation_files = lead.get("conversation_files", [])
                if conversation_files:
                    # Return the most recent file (last in list)
                    return conversation_files[-1]
            
            # Also check if any conversation file matches the client_id pattern
            conversation_files = lead.get("conversation_files", [])
            for conv_file in conversation_files:
                if client_id in conv_file:
                    return conv_file
        
        # If not found in database, check filesystem for existing files from today
        if os.path.exists(EXCEL_FOLDER):
            today = datetime.now().strftime('%Y-%m-%d')
            matching_files = []
            for filename in os.listdir(EXCEL_FOLDER):
                if (filename.startswith(f"conversation_{client_id}_") and 
                    filename.endswith(".txt") and 
                    today in filename):
                    matching_files.append(filename)
            
            # Return the most recent file from today if any found
            if matching_files:
                matching_files.sort()  # Sort by timestamp in filename
                return matching_files[-1]
        
        return None
        
    except Exception as e:
        logging.error(f"Error finding existing conversation file: {e}")
        return None

def get_or_create_client_and_conversation(phone: str = None, email: str = None, client_id: str = None, name: str = None) -> dict:
    """
    Always returns a client dict (existing or new) and the conversation file to use.
    - Checks by phone, email, or conversation file (client_id)
    - If found, returns existing client and latest conversation file
    - If not found, creates new client, assigns new ID, and starts new conversation file
    """
    leads = load_leads_minimal()
    client = find_client_by_contact(phone, email, client_id)
    created = False
    if client:
        # Use latest conversation file if exists, else create new
        conversation_files = client.get("conversation_files", [])
        if conversation_files:
            conversation_file = conversation_files[-1]
        else:
            # Create new conversation file for this client
            timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
            conversation_file = f"conversation_{client.get('phone','') or client_id or 'unknown'}_{timestamp}.txt"
            client.setdefault("conversation_files", []).append(conversation_file)
            
            # Create the actual file on disk
            file_path = f"client_data/{conversation_file}"
            with open(file_path, 'w', encoding='utf-8') as f:
                f.write(f"# Conversation started: {datetime.now().isoformat()}\n")
                f.write(f"# Client ID: {client['id']}\n")
                f.write(f"# Phone: {client.get('phone', 'N/A')}\n")
                f.write(f"# Email: {client.get('email', 'N/A')}\n\n")
            
            save_leads_minimal(leads)
    else:
        # Create new client
        new_id = max([l.get("id", 0) for l in leads] + [0]) + 1
        timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        conversation_file = f"conversation_{phone or client_id or 'unknown'}_{timestamp}.txt"
        
        # Create the actual file on disk
        file_path = f"client_data/{conversation_file}"
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(f"# Conversation started: {datetime.now().isoformat()}\n")
            f.write(f"# Client ID: {new_id}\n")
            f.write(f"# Phone: {phone or 'N/A'}\n")
            f.write(f"# Email: {email or 'N/A'}\n\n")
        
        client = {
            "id": new_id,
            "name": name or "",
            "phone": phone or "",
            "email": email or "",
            "chat_summary": "Conversation in progress",
            "meet_link": "",
            "meeting_date": "",
            "meeting_time": "",
            "meeting_timezone": "",
            "conversation_files": [conversation_file],
            "last_interaction": datetime.now().isoformat(),
            "total_messages": 0
        }
        leads.append(client)
        save_leads_minimal(leads)
        created = True
    return {"client": client, "conversation_file": conversation_file, "created": created}


def saba_greeting_with_client_id(client: dict) -> str:
    """
    Returns Saba's greeting message including the client ID.
    """
    client_id = str(client.get('id', ''))  # Convert to string to handle both int and str IDs
    return f"Hi! I'm Saba from IMJD. Your client ID is {client_id}. Please use this for future reference. How can I help you today?"


# 🎯 CLIENT DEDUPLICATION API ENDPOINTS
@app.route('/api/get_or_create_client', methods=['POST'])
@limiter.limit("1000 per hour")
def api_get_or_create_client():
    """
    API endpoint to get or create client using the deduplication utility
    """
    try:
        data = request.get_json()
        phone = data.get('phone')
        email = data.get('email')
        client_id = data.get('client_id')
        conversation_file = data.get('conversation_file')
        
        # Use the utility function
        result = get_or_create_client_and_conversation(phone, email, client_id, conversation_file)
        
        return jsonify(result)
    except Exception as e:
        logging.error(f"Error in get_or_create_client API: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/saba_greeting', methods=['POST'])
@limiter.limit("1000 per hour")
def api_saba_greeting():
    """
    API endpoint to get Saba's greeting with client ID
    """
    try:
        data = request.get_json()
        client = data.get('client')
        
        if not client:
            return jsonify({"error": "Client data required"}), 400
        
        # Use the utility function
        greeting = saba_greeting_with_client_id(client)
        
        return jsonify({"greeting": greeting})
    except Exception as e:
        logging.error(f"Error in saba_greeting API: {e}")
        return jsonify({"error": str(e)}), 500

if __name__ == '__main__':
    port = 5000
    if len(sys.argv) > 1:
        try:
            port = int(sys.argv[1])
        except ValueError:
            print("Invalid port number. Using default port 5000.")
    app.run(debug=True, port=port)